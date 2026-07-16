"""
Generate mirae_dcf.xlsx — a live DCF workbook for Mirae Corporation (025560).

    python3 build_dcf_xlsx.py            # -> mirae_dcf.xlsx
    python3 build_dcf_xlsx.py --out /tmp/x.xlsx

Design: the workbook is GENERATED, never hand-maintained. CONTRACTS/ACTUALS/CFG
are read live from index.html via mirae_data.py, so the sheet can never drift
from the dashboard. Re-run after any new filing.

Formulas live in the cells — this is a model you drive from Excel, not a value
dump. Every input is on the Assumptions sheet (blue). Everything else computes.

Sources: FY25 사업보고서 rcept 20260319001166 (consolidated BS/CF + note 30 tax),
contract filings per OrderBook sheet. See README sheet for what is measured vs
assumed.
"""

import argparse
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import mirae_data

REPO = os.path.dirname(os.path.abspath(__file__))

# ---- FY2025 actuals, hand-verified from the DART consolidated statements ----
# (KRW bn. Kept here, not in index.html, because the dashboard has no use for them.)
FY25 = {
    "cash": 10.87, "ar": 12.08, "inventory": 34.64, "ap": 9.37,
    "st_debt": 7.45, "convertible": 7.46, "lease": 0.91,
    "invest_property": 53.25, "ppe": 16.99,
    "total_assets": 152.35, "total_liab": 29.75, "equity": 122.60,
    "da": 0.817, "amort": 0.024, "capex": 0.084,
    "revenue": 50.78, "pretax": 9.83, "current_tax": 0.0223,
}
# note 30-(5) expiry schedule, measured at 2025-12-31 (KRW bn)
NOL_BUCKETS = [("2026", 2.45), ("2027", 27.10), ("2028", 10.46), ("2029+", 25.39)]
TAX_CREDITS = 4.61
STATUTORY = 0.209          # 2.05bn / 9.83bn per note 30-(2)
YEARS = [2026, 2027, 2028, 2029, 2030]

# ---- styling ----
H1 = Font(bold=True, size=14)
H2 = Font(bold=True, size=11, color="FFFFFF")
BOLD = Font(bold=True)
INPUT = Font(color="0000CC", bold=True)          # blue = editable
NOTE = Font(italic=True, size=9, color="666666")
FILL_H = PatternFill("solid", fgColor="1F3864")
FILL_IN = PatternFill("solid", fgColor="DDEBF7")
FILL_TOT = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
NUM = "#,##0.00"
PCT = "0.0%"


def hdr(ws, row, text, span=8):
    c = ws.cell(row, 1, text)
    c.font = H2
    for i in range(1, span + 1):
        ws.cell(row, i).fill = FILL_H
    return row + 1


def put(ws, row, label, value=None, fmt=NUM, font=None, note=None):
    ws.cell(row, 1, label)
    if font:
        ws.cell(row, 1).font = font
    if value is not None:
        c = ws.cell(row, 2, value)
        c.number_format = fmt
    if note:
        n = ws.cell(row, 4, note)
        n.font = NOTE
    return row + 1


def build(path):
    wb = Workbook()

    # =========================== README ===========================
    ws = wb.active
    ws.title = "README"
    ws.column_dimensions["A"].width = 118
    r = 1
    ws.cell(r, 1, "Mirae Corporation (025560) — DCF").font = H1
    r += 2
    for line in [
        "GENERATED FILE — do not hand-edit. Re-run build_dcf_xlsx.py after any new DART filing.",
        "Contracts and quarterly actuals are read live from index.html, so this workbook cannot drift from the dashboard.",
        "",
        "Edit only the BLUE cells on the Assumptions sheet. Everything else is a formula.",
        "",
        "WHAT IS MEASURED (from DART, FY25 사업보고서 rcept 20260319001166):",
        "   Order book: 22 contracts, reconciled against all 29 단일판매·공급계약 filings. Total 115.48bn.",
        "   Quarterly segment revenue Q1'25..Q1'26; FY25 consolidated BS + CF; note 30 tax detail.",
        "",
        "WHAT IS ASSUMED (and therefore where the answer actually comes from):",
        "   FY2027+ revenue growth. The order book runs dry after Q3'27 — only 3.58bn is contracted beyond FY26.",
        "   WACC and terminal growth. Terminal value will dominate; see the Sensitivity sheet before believing any",
        "   single fair-value number. With 5 quarters of actuals on a company that just 5x'd quarterly revenue,",
        "   this model is a disciplined way to express assumptions — not a measurement of intrinsic value.",
        "",
        "THREE THINGS THAT DRIVE THIS MODEL MORE THAN THE REVENUE LINE:",
        "",
        "1. WORKING CAPITAL. FY25 NWC = AR 12.08 + Inv 34.64 - AP 9.37 = 37.35bn on 50.78bn revenue (~74%).",
        "   FY24 was 23.26bn, so dWC consumed 14.1bn of cash in a year Mirae earned 9.83bn pretax — cash FELL",
        "   13.50 -> 10.87. Inventory nearly doubled. If NWC stays ~74% of revenue, a revenue ramp to ~115bn implies",
        "   a further ~48bn WC build, which would swamp EBIT and make FCF deeply negative. THIS is the swing factor.",
        "   The NWC% input is the single most consequential cell in the workbook.",
        "",
        "2. TAX. Mirae has 65.4bn of unused tax losses and pays ~no cash tax (FY25 current tax 22.3m on 9.83bn",
        "   pretax). SME status holds a 100% offset (법인세법 제13조①), and even a threshold breach buys a 7-year",
        "   grace period (KOSPI-listed). So the cap is not the constraint — EXPIRY is: 29.6bn dies within 2 years.",
        "   Modelled as an expiry-ordered waterfall on the NOL sheet, not a flat rate.",
        "",
        "3. NON-OPERATING ASSETS. 53.25bn of 투자부동산 (investment property) — 35% of total assets, reclassified",
        "   from PP&E in FY25. It generates rent, not equipment revenue, so it is added to EV as a non-operating",
        "   asset at book rather than being driven by the forecast. Book value is a floor-ish proxy, not a valuation.",
        "",
        "KNOWN GAPS — read before relying on this:",
        "   - Only FY25/FY24 balance sheet. No quarterly WC history, so the WC model is anchored on ONE year of",
        "     delta. That is thin for the variable that matters most.",
        "   - 7.46bn of 유동성전환사채 (convertible bonds) sits in net debt but its dilution is NOT in the share",
        "     count. Conversion terms not yet pulled. Equity value per share is overstated if they convert.",
        "   - Share count 27,778,678 assumes the 2,324,538-share placement closes 2026-07-27. It had not as of",
        "     2026-07-16 and has been pushed twice.",
        "   - 계약기간 종료일 is a deliver-BY deadline, not a delivery date. Mirae ships early (Q1'26 booked 20.7bn",
        "     of revenue against ZERO order-book deadlines), so delivery-year bucketing systematically LAGS revenue.",
        "   - Cost of equity / WACC is an input, not derived. No beta study was done.",
    ]:
        ws.cell(r, 1, line)
        r += 1

    # ========================= ASSUMPTIONS =========================
    ws = wb.create_sheet("Assumptions")
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["D"].width = 74
    r = hdr(ws, 1, "ASSUMPTIONS — edit blue cells only", 5)
    r += 1
    rows = {}

    def inp(row, key, label, val, fmt=NUM, note=None):
        ws.cell(row, 1, label)
        c = ws.cell(row, 2, val)
        c.font = INPUT
        c.fill = FILL_IN
        c.border = BOX
        c.number_format = fmt
        if note:
            ws.cell(row, 4, note).font = NOTE
        rows[key] = row
        return row + 1

    r = put(ws, r, "OPERATING", font=BOLD)
    r = inp(r, "baseAte", "Base / sub-threshold ATE (bn/qtr)", 5.0, NUM,
            "Not disclosed. Ran ~4.6bn/qtr in H1'25 before any big contract delivered.")
    r = inp(r, "mai", "MAI revenue (bn/qtr)", 2.0, NUM,
            "Volatile 0.85-3.39, avg ~1.8. Model absolute, NOT % of revenue — it is independent of ATE.")
    r = inp(r, "gm", "Gross margin %", 0.42, PCT,
            "Actuals swing 34.6-56.7% with mix. FY25 blended ~41%.")
    r = inp(r, "sga", "SG&A (bn/qtr)", 2.4, NUM,
            "~FIXED: ran 2.3bn on both a 4.75bn and a 22.6bn revenue quarter. Operating leverage is the thesis.")
    r = inp(r, "rnd", "R&D (bn/qtr)", 0.12, NUM, "Also ~fixed, small.")
    r = inp(r, "growth", "Revenue growth FY2027+ %", 0.10, PCT,
            "PURE ASSUMPTION. Order book is dry after Q3'27 (only 3.58bn contracted). Drives most of the answer.")
    r += 1

    r = put(ws, r, "WORKING CAPITAL  <-- the most consequential input", font=BOLD)
    r = inp(r, "nwc", "NWC as % of revenue", 0.735, PCT,
            "FY25 actual: 37.35/50.78 = 73.6%. Set lower only if you believe the ramp brings WC efficiency.")
    r += 1

    r = put(ws, r, "CAPITAL / VALUATION", font=BOLD)
    r = inp(r, "capex", "Capex (bn/yr)", 0.30, NUM,
            "FY25 actual only 0.084bn (<0.2% of revenue). Asset-light. Raised slightly for a ramp.")
    r = inp(r, "da", "D&A (bn/yr)", 0.85, NUM, "FY25: 0.817 depreciation + 0.024 amortisation.")
    r = inp(r, "wacc", "WACC", 0.12, PCT, "INPUT, not derived. No beta study done. See Sensitivity.")
    r = inp(r, "tg", "Terminal growth", 0.02, PCT, "TV will dominate EV — check the Sensitivity sheet.")
    r += 1

    r = put(ws, r, "TAX", font=BOLD)
    r = inp(r, "taxrate", "Statutory rate", STATUTORY, PCT, "2.05bn / 9.83bn per note 30-(2).")
    r = inp(r, "nolpct", "NOL offset % of taxable income", 1.00, PCT,
            "100% = SME (조특법 제6조①). 80% only if 나목 breached (>=500bn-asset holder takes >=30%) — no grace.")
    r = inp(r, "mintax", "Minimum tax (최저한세) floor %", 0.07, PCT,
            "SME rate. Floors CREDIT usage, not NOL usage. SECONDARY-sourced — verify before relying on it.")
    r += 1

    r = put(ws, r, "CAPITAL STRUCTURE", font=BOLD)
    r = inp(r, "shares", "Shares (M, fully diluted)", 27.778678, "#,##0.000000",
            "Assumes the 2.32M placement closes 2026-07-27. Excludes convertible-bond dilution.")
    r = inp(r, "netdebt", "Net debt (bn)", round(FY25["st_debt"] + FY25["convertible"] + FY25["lease"] - FY25["cash"], 2),
            NUM, "Debt 7.45 + CB 7.46 + lease 0.91 - cash 10.87. CB dilution NOT in share count.")
    r = inp(r, "invprop", "Investment property (bn)", FY25["invest_property"], NUM,
            "Non-operating; added to EV at book. 35% of total assets.")
    A = {k: "Assumptions!$B$%d" % v for k, v in rows.items()}

    # ========================== ORDER BOOK ==========================
    ws = wb.create_sheet("OrderBook")
    contracts, actuals, cfg = mirae_data.load()
    heads = ["Order date", "Counterparty", "Segment", "Memory", "Value (bn)", "Delivery", "Delivery Q", "Delivery FY", "Terms"]
    for i, h in enumerate(heads, 1):
        c = ws.cell(1, i, h)
        c.font = H2
        c.fill = FILL_H
    for j, c in enumerate(sorted(contracts, key=lambda x: x["date"]), 2):
        ws.cell(j, 1, c["date"]); ws.cell(j, 2, c["party"]); ws.cell(j, 3, c["seg"])
        ws.cell(j, 4, c["mem"])
        ws.cell(j, 5, c["value"]).number_format = NUM
        ws.cell(j, 6, c["delivery"])
        ws.cell(j, 7, mirae_data.quarter_of(c["delivery"]))
        ws.cell(j, 8, int(c["delivery"][:4]))
        ws.cell(j, 9, c.get("terms", ""))
    last = len(contracts) + 1
    tr = last + 1
    ws.cell(tr, 4, "TOTAL").font = BOLD
    ws.cell(tr, 5, "=SUM(E2:E%d)" % last).number_format = NUM
    ws.cell(tr, 5).font = BOLD
    for w, col in zip([12, 30, 20, 9, 11, 12, 11, 11, 74], "ABCDEFGHI"):
        ws.column_dimensions[col].width = w
    ws.cell(tr + 2, 1, "Reconciled against all 29 단일판매·공급계약 filings (2025-01..2026-07); 29 -> 22 distinct after "
                       "deduping [기재정정] revisions. Delivery = 계약기간 종료일 = deliver-BY deadline, not actual "
                       "delivery — Mirae ships early, so this LAGS revenue.").font = NOTE

    # =========================== ACTUALS ===========================
    ws = wb.create_sheet("Actuals")
    ws.column_dimensions["A"].width = 30
    r = hdr(ws, 1, "QUARTERLY ACTUALS (KRW bn) — DART 매출실적 + stockanalysis margins", 8)
    qs = sorted(actuals)
    for i, q in enumerate(qs, 2):
        ws.cell(r, i, q).font = BOLD
    r += 1
    for key, lab in [("ate", "ATE revenue"), ("mai", "MAI revenue"), ("rev", "Total revenue"),
                     ("gm", "Gross margin %"), ("sga", "SG&A"), ("rnd", "R&D"), ("op", "Operating profit")]:
        ws.cell(r, 1, lab)
        for i, q in enumerate(qs, 2):
            ws.cell(r, i, actuals[q][key]).number_format = NUM
        r += 1
    r += 1
    r = hdr(ws, r, "FY2025 BALANCE SHEET / CASH FLOW (consolidated, rcept 20260319001166)", 8)
    for lab, key in [("Cash", "cash"), ("Receivables", "ar"), ("Inventory", "inventory"),
                     ("Payables", "ap"), ("Short-term debt", "st_debt"),
                     ("Convertible bonds", "convertible"), ("Lease liabilities", "lease"),
                     ("Investment property", "invest_property"), ("PP&E", "ppe"),
                     ("Total assets", "total_assets"), ("Equity", "equity"),
                     ("Depreciation", "da"), ("Amortisation", "amort"), ("Capex", "capex")]:
        r = put(ws, r, lab, FY25[key])
    r += 1
    ws.cell(r, 1, "NWC (AR + Inventory - AP)").font = BOLD
    ws.cell(r, 2, round(FY25["ar"] + FY25["inventory"] - FY25["ap"], 2)).number_format = NUM
    r += 1
    ws.cell(r, 1, "NWC % of revenue").font = BOLD
    ws.cell(r, 2, (FY25["ar"] + FY25["inventory"] - FY25["ap"]) / FY25["revenue"]).number_format = PCT
    r += 1
    ws.cell(r, 1, "FY24 NWC was 23.26bn -> dWC consumed 14.1bn of cash in FY25, vs 9.83bn pretax income. "
                  "Cash fell 13.50 -> 10.87.").font = NOTE

    # ============================= NOL =============================
    ws = wb.create_sheet("NOL")
    ws.column_dimensions["A"].width = 34
    r = hdr(ws, 1, "TAX LOSS WATERFALL — expiry-ordered (note 30-(4)/(5), at 2025-12-31)", 8)
    r += 1
    ws.cell(r, 1, "Expiry bucket").font = BOLD
    ws.cell(r, 2, "Amount (bn)").font = BOLD
    br = {}
    for i, (lab, amt) in enumerate(NOL_BUCKETS):
        r += 1
        ws.cell(r, 1, "Expires " + lab)
        ws.cell(r, 2, amt).number_format = NUM
        br[lab] = r
    r += 1
    ws.cell(r, 1, "TOTAL").font = BOLD
    ws.cell(r, 2, "=SUM(B%d:B%d)" % (br["2026"], br["2029+"])).number_format = NUM
    ws.cell(r, 2).font = BOLD
    r += 2
    ws.cell(r, 1, "29.55bn (45%) expires by end-2027. Use-it-or-lose-it: only taxable income earned NOW saves it. "
                  "This — not the 80/100% cap — is the binding constraint.").font = NOTE
    r += 2

    r = hdr(ws, r, "CONSUMPTION BY YEAR", 8)
    hrow = r
    for i, y in enumerate(YEARS, 2):
        ws.cell(hrow, i, y).font = BOLD
    r += 1
    rowmap = {}
    for lab in ["Taxable income (pre-NOL)", "NOL available", "NOL used", "NOL expired unused",
                "NOL closing balance", "Taxable after NOL",
                "Tax at statutory", "Minimum tax floor",
                "Credits available", "Credits used", "Credits closing"]:
        ws.cell(r, 1, lab)
        rowmap[lab] = r
        r += 1
    r += 1
    ws.cell(r, 1, "Credits (4.61bn) offset tax only down to the 최저한세 floor — they cannot take cash tax to zero. "
                  "The floor applies to the base AFTER the NOL deduction (이월결손금 is a 법인세법 제13조 deduction, "
                  "not a 조특법 특례, so it is not added back).").font = NOTE
    return wb, ws, A, rowmap, hrow, br, contracts, cfg


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=os.path.join(REPO, "mirae_dcf.xlsx"))
    args = ap.parse_args()
    wb, ws, A, rowmap, hrow, br, contracts, cfg = build(args.out)
    finish(wb, ws, A, rowmap, hrow, br, contracts, cfg)
    wb.save(args.out)
    print("wrote %s" % args.out)


def finish(wb, nol, A, rm, hrow, br, contracts, cfg):
    """NOL formulas + Model + Sensitivity. Split out to keep build() readable."""
    # order book by delivery FY
    obk = {}
    for c in contracts:
        obk[int(c["delivery"][:4])] = round(obk.get(int(c["delivery"][:4]), 0) + c["value"], 4)

    M = wb.create_sheet("Model")
    M.column_dimensions["A"].width = 34
    r = hdr(M, 1, "DCF MODEL (KRW bn) — all figures annual", 8)
    for i, y in enumerate(YEARS, 2):
        M.cell(r, i, y).font = BOLD
    r += 1
    R = {}

    def line(label, f, fmt=NUM, bold=False, note=None):
        nonlocal r
        R[label] = r            # register BEFORE evaluating f — Revenue references its own prior column
        M.cell(r, 1, label)
        if bold:
            M.cell(r, 1).font = BOLD
        for i, y in enumerate(YEARS, 2):
            col = get_column_letter(i)
            c = M.cell(r, i, f(y, col, i))
            c.number_format = fmt
            if bold:
                c.font = BOLD
                c.fill = FILL_TOT
        if note:
            M.cell(r, 8, note).font = NOTE
        R[label] = r
        r += 1

    line("Order book (contracted)", lambda y, col, i: obk.get(y, 0.0), note="Dry after FY2027 — 3.58bn only.")
    line("Base ATE", lambda y, col, i: "=%s*4" % A["baseAte"])
    line("MAI", lambda y, col, i: "=%s*4" % A["mai"])
    line("Revenue", lambda y, col, i:
         ("=SUM(%s%d:%s%d)" % (col, R["Order book (contracted)"], col, R["MAI"])) if y == 2026
         else ("=%s%d*(1+%s)" % (get_column_letter(i - 1), R["Revenue"], A["growth"])),
         bold=True, note="FY26 = contracted + base. FY27+ = growth assumption.")
    line("Gross profit", lambda y, col, i: "=%s%d*%s" % (col, R["Revenue"], A["gm"]))
    line("SG&A", lambda y, col, i: "=%s*4" % A["sga"])
    line("R&D", lambda y, col, i: "=%s*4" % A["rnd"])
    line("EBIT", lambda y, col, i: "=%s%d-%s%d-%s%d" % (col, R["Gross profit"], col, R["SG&A"], col, R["R&D"]), bold=True)
    line("EBIT margin", lambda y, col, i: "=IFERROR(%s%d/%s%d,0)" % (col, R["EBIT"], col, R["Revenue"]), fmt=PCT)
    r += 1

    # ---- NOL waterfall on the NOL sheet, driven by Model EBIT ----
    for i, y in enumerate(YEARS, 2):
        col = get_column_letter(i)
        prev = get_column_letter(i - 1)
        nol.cell(rm["Taxable income (pre-NOL)"], i, "=Model!%s%d" % (col, R["EBIT"])).number_format = NUM
        if y == 2026:
            avail = "=SUM(B%d:B%d)" % (br["2026"], br["2029+"])
        else:
            avail = "=%s%d" % (prev, rm["NOL closing balance"])
        nol.cell(rm["NOL available"], i, avail).number_format = NUM
        nol.cell(rm["NOL used"], i, "=MAX(0,MIN(%s%d,%s%d*%s))" % (
            col, rm["NOL available"], col, rm["Taxable income (pre-NOL)"], A["nolpct"])).number_format = NUM
        # expire whatever sits in the bucket that dies this year and wasn't used
        exp = {2026: br["2026"], 2027: br["2027"], 2028: br["2028"]}.get(y)
        nol.cell(rm["NOL expired unused"], i,
                 ("=MAX(0,B%d-%s%d)" % (exp, col, rm["NOL used"])) if exp else 0).number_format = NUM
        nol.cell(rm["NOL closing balance"], i, "=MAX(0,%s%d-%s%d-%s%d)" % (
            col, rm["NOL available"], col, rm["NOL used"], col, rm["NOL expired unused"])).number_format = NUM
        nol.cell(rm["Taxable after NOL"], i, "=MAX(0,%s%d-%s%d)" % (
            col, rm["Taxable income (pre-NOL)"], col, rm["NOL used"])).number_format = NUM
        # --- tax + credit pool. Credits offset down to the 최저한세 floor, never below. ---
        nol.cell(rm["Tax at statutory"], i, "=%s%d*%s" % (
            col, rm["Taxable after NOL"], A["taxrate"])).number_format = NUM
        nol.cell(rm["Minimum tax floor"], i, "=%s%d*%s" % (
            col, rm["Taxable after NOL"], A["mintax"])).number_format = NUM
        nol.cell(rm["Credits available"], i,
                 (TAX_CREDITS if y == 2026 else "=%s%d" % (prev, rm["Credits closing"]))).number_format = NUM
        nol.cell(rm["Credits used"], i, "=MAX(0,MIN(%s%d,%s%d-%s%d))" % (
            col, rm["Credits available"], col, rm["Tax at statutory"],
            col, rm["Minimum tax floor"])).number_format = NUM
        nol.cell(rm["Credits closing"], i, "=%s%d-%s%d" % (
            col, rm["Credits available"], col, rm["Credits used"])).number_format = NUM
    nol.cell(rm["NOL expired unused"], 8,
             "Simplification: expires the whole dated bucket less that year's usage. With income this large the "
             "buckets drain first, so this is conservative-neutral; revisit if EBIT assumptions fall a lot.").font = NOTE

    line("Taxable income", lambda y, col, i: "=NOL!%s%d" % (col, rm["Taxable income (pre-NOL)"]))
    line("less NOL used", lambda y, col, i: "=-NOL!%s%d" % (col, rm["NOL used"]))
    line("Taxable after NOL", lambda y, col, i: "=NOL!%s%d" % (col, rm["Taxable after NOL"]))
    line("Tax at statutory", lambda y, col, i: "=NOL!%s%d" % (col, rm["Tax at statutory"]))
    line("Minimum tax floor (최저한세)", lambda y, col, i: "=NOL!%s%d" % (col, rm["Minimum tax floor"]),
         note="A FLOOR on credit usage, not a cap on tax. Applies to the base AFTER the NOL deduction.")
    line("less credits used", lambda y, col, i: "=-NOL!%s%d" % (col, rm["Credits used"]),
         note="4.61bn pool; offsets tax only down to the floor.")
    line("Cash tax", lambda y, col, i: "=MAX(%s%d-NOL!%s%d,0)" % (
        col, R["Tax at statutory"], col, rm["Credits used"]), bold=True)
    line("NOPAT", lambda y, col, i: "=%s%d-%s%d" % (col, R["EBIT"], col, R["Cash tax"]), bold=True)
    r += 1

    line("add D&A", lambda y, col, i: "=%s" % A["da"])
    line("less Capex", lambda y, col, i: "=-%s" % A["capex"])
    line("NWC", lambda y, col, i: "=%s%d*%s" % (col, R["Revenue"], A["nwc"]))
    line("less change in NWC", lambda y, col, i:
         ("=-(%s%d-37.35)" % (col, R["NWC"])) if y == 2026
         else ("=-(%s%d-%s%d)" % (col, R["NWC"], get_column_letter(i - 1), R["NWC"])),
         note="FY25 actual NWC = 37.35bn is the opening base. THE swing factor — see README.")
    line("Free cash flow", lambda y, col, i: "=%s%d+%s%d+%s%d+%s%d" % (
        col, R["NOPAT"], col, R["add D&A"], col, R["less Capex"], col, R["less change in NWC"]), bold=True)
    line("Discount factor", lambda y, col, i: "=1/(1+%s)^%d" % (A["wacc"], y - 2025), fmt="0.0000")
    line("PV of FCF", lambda y, col, i: "=%s%d*%s%d" % (col, R["Free cash flow"], col, R["Discount factor"]), bold=True)
    r += 1

    # ---- valuation bridge ----
    r = hdr(M, r, "VALUATION BRIDGE", 8)
    V = {}

    def v(label, formula, fmt=NUM, bold=False, note=None):
        nonlocal r
        M.cell(r, 1, label)
        c = M.cell(r, 2, formula)
        c.number_format = fmt
        if bold:
            M.cell(r, 1).font = BOLD
            c.font = BOLD
            c.fill = FILL_TOT
        if note:
            M.cell(r, 4, note).font = NOTE
        V[label] = r
        r += 1

    lastc = get_column_letter(1 + len(YEARS))
    v("Sum PV of explicit FCF", "=SUM(B%d:%s%d)" % (R["PV of FCF"], lastc, R["PV of FCF"]))
    v("Terminal value (Gordon)", "=%s%d*(1+%s)/(%s-%s)" % (lastc, R["Free cash flow"], A["tg"], A["wacc"], A["tg"]),
      note="Breaks down if terminal FCF is negative — check the FCF line before trusting this.")
    v("PV of terminal value", "=B%d*%s%d" % (V["Terminal value (Gordon)"], lastc, R["Discount factor"]))
    v("TV as % of EV", "=IFERROR(B%d/(B%d+B%d),0)" % (V["PV of terminal value"], V["Sum PV of explicit FCF"],
                                                      V["PV of terminal value"]), fmt=PCT,
      note="If this is >80%, the DCF is an exit-multiple guess wearing a cash-flow costume.")
    v("Enterprise value", "=B%d+B%d" % (V["Sum PV of explicit FCF"], V["PV of terminal value"]), bold=True)
    v("add Investment property", "=%s" % A["invprop"], note="Non-operating, at book. Not a valuation.")
    v("less Net debt", "=-%s" % A["netdebt"])
    v("Equity value", "=B%d+B%d+B%d" % (V["Enterprise value"], V["add Investment property"], V["less Net debt"]),
      bold=True)
    v("Shares (M)", "=%s" % A["shares"], fmt="#,##0.000000")
    v("Value per share (KRW)", "=IFERROR(B%d*1000/B%d,0)" % (V["Equity value"], V["Shares (M)"]), fmt="#,##0",
      bold=True, note="Excludes convertible-bond dilution (7.46bn CB outstanding).")
    v("Last traded price (KRW)", 9140, fmt="#,##0",
      note="45,700 on 2026-07-03 / 5 for the split. Stock SUSPENDED 07-07..07-24; not a live quote.")
    v("Upside / (downside)", "=IFERROR(B%d/B%d-1,0)" % (V["Value per share (KRW)"], V["Last traded price (KRW)"]),
      fmt=PCT)

    # ========================= SENSITIVITY =========================
    S = wb.create_sheet("Sensitivity")
    S.column_dimensions["A"].width = 30
    r = hdr(S, 1, "SENSITIVITY — value per share (KRW)", 8)
    r += 1
    S.cell(r, 1, "Excel recalculates these only via Data > What-If > Data Table, or just change the "
                 "Assumptions inputs and watch Model!B%d." % V["Value per share (KRW)"]).font = NOTE
    r += 2
    S.cell(r, 1, "The grid below is a MANUAL reference: set WACC / terminal growth on Assumptions and read the "
                 "result. Automating it needs a Data Table, which openpyxl cannot emit.").font = NOTE
    r += 2
    S.cell(r, 1, "WACC \\ terminal g").font = BOLD
    for j, g in enumerate([0.00, 0.01, 0.02, 0.03, 0.04], 2):
        S.cell(r, j, g).number_format = PCT
        S.cell(r, j).font = BOLD
    for i, w in enumerate([0.08, 0.10, 0.12, 0.14, 0.16], 1):
        S.cell(r + i, 1, w).number_format = PCT
        S.cell(r + i, 1).font = BOLD
        for j in range(2, 7):
            S.cell(r + i, j, "—").alignment = Alignment(horizontal="center")
    r += 7
    S.cell(r, 1, "Terminal value will likely be most of EV. Before reading any number above as a fair value, check "
                 "Model!B%d (TV as %% of EV) and confirm terminal FCF is positive — if WC is still building at the "
                 "terminal year, Gordon growth on a negative FCF produces a meaningless negative TV." %
           V["TV as % of EV"]).font = NOTE


if __name__ == "__main__":
    main()
